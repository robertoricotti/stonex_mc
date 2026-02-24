import os
import math
import uuid
import csv
import shutil
import sys
from datetime import datetime, timezone

import tkinter as tk
from tkinter import filedialog, messagebox

from tkinterdnd2 import DND_FILES, TkinterDnD
import openpyxl
import xml.etree.ElementTree as ET


# ----------------------------
# Namespaces
# ----------------------------
NS_DRP = "http://www.iredes.org/xml/DrillRig"
NS_IR  = "http://www.iredes.org/xml"
NS_JET = "urn:custom:jet:treatment:v1"   # DEVE combaciare con Android parser

ET.register_namespace("", NS_DRP)
ET.register_namespace("IR", NS_IR)
ET.register_namespace("JET", NS_JET)


def _is_empty(v) -> bool:
    if v is None:
        return True
    if isinstance(v, str) and v.strip() == "":
        return True
    return False


def desktop_path() -> str:
    # Windows: tipicamente C:\Users\...\Desktop
    # (se non esiste, ripiega alla home)
    home = os.path.expanduser("~")
    cand = os.path.join(home, "Desktop")
    return cand if os.path.isdir(cand) else home

def app_dir() -> str:
    """
    Ritorna la cartella base dell'app.
    - In sviluppo: la cartella dove sta lo script .py
    - In exe (PyInstaller): la cartella dell'eseguibile
    """
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


def resource_path(relative: str) -> str:
    """
    Se impacchetti con PyInstaller, i file inclusi con --add-data finiscono in sys._MEIPASS.
    In sviluppo, stanno accanto allo script.
    """
    base = getattr(sys, "_MEIPASS", app_dir())
    return os.path.join(base, relative)


def copy_template_if_missing(src_path: str, dst_folder: str, dst_name: str):
    os.makedirs(dst_folder, exist_ok=True)
    dst = os.path.join(dst_folder, dst_name)
    if os.path.exists(dst):
        return  # non sovrascrivere
    shutil.copyfile(src_path, dst)


def ensure_project_folders() -> dict:
    base = os.path.join(desktop_path(), "STX Hole Converter")
    rock = os.path.join(base, "ROCK Proj")
    jet  = os.path.join(base, "JET Proj")
    solar= os.path.join(base, "SOLAR Proj")

    os.makedirs(rock, exist_ok=True)
    os.makedirs(jet, exist_ok=True)
    os.makedirs(solar, exist_ok=True)

    # --- Copia templates vuoti dentro le cartelle progetto (solo se mancano) ---
    # I file template devono essere inclusi accanto allo script, oppure come risorse PyInstaller.
    copy_template_if_missing(resource_path("templates/Jet_Template.xlsx"),   jet,   "Jet_Template.xlsx")
    copy_template_if_missing(resource_path("templates/Rock_Template.xlsx"),  rock,  "Rock_Template.xlsx")
    copy_template_if_missing(resource_path("templates/Solar_Template.xlsx"), solar, "Solar_Template.xlsx")

    return {"base": base, "ROCK": rock, "JET": jet, "SOLAR": solar}



def open_in_explorer(path: str):
    # Windows: os.startfile apre Explorer
    try:
        os.startfile(path)
    except Exception:
        pass


def compute_geometry(xh, yh, zh, xe, ye, ze):
    dx = xe - xh
    dy = ye - yh
    dz = ze - zh

    horiz = math.hypot(dx, dy)
    length = math.sqrt(dx * dx + dy * dy + dz * dz)

    # Bearing: 0°=Nord, 90°=Est
    bearing = math.degrees(math.atan2(dx, dy))
    if bearing < 0:
        bearing += 360.0

    # Tilt (come nel tuo Android): 0 verticale, 90 orizzontale
    vert = abs(dz)
    tilt = math.degrees(math.atan2(horiz, vert)) if not (horiz == 0 and vert == 0) else 0.0

    return {
        "dx": dx, "dy": dy, "dz": dz,
        "horiz": horiz,
        "length": length,
        "bearing_deg": bearing,
        "tilt_deg": tilt,
    }


# ----------------------------
# Readers (Excel/CSV)
# ----------------------------
def read_table_from_excel(path: str):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    headers = [("" if h is None else str(h).strip()) for h in headers]

    rows = []
    for r in range(2, ws.max_row + 1):
        row = {}
        empty_row = True
        for c, h in enumerate(headers, start=1):
            if h == "":
                continue
            v = ws.cell(r, c).value
            if not _is_empty(v):
                empty_row = False
            row[h] = v
        if not empty_row:
            rows.append(row)
    return headers, rows


def read_table_from_csv(path: str):
    # Prova delimitatori comuni: ; o ,
    # (Se vuoi, lo rendiamo opzione in UI)
    def try_read(delim):
        with open(path, "r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f, delimiter=delim)
            if reader.fieldnames is None:
                return None
            rows = list(reader)
            headers = [h.strip() for h in reader.fieldnames]
            norm_rows = []
            for rr in rows:
                row = {}
                empty_row = True
                for h in headers:
                    v = rr.get(h, None)
                    if v is not None:
                        v = v.strip()
                    if v != "":
                        empty_row = False
                    row[h] = (None if v == "" else v)
                if not empty_row:
                    norm_rows.append(row)
            return headers, norm_rows

    out = try_read(";")
    if out is not None:
        return out
    out = try_read(",")
    if out is not None:
        return out
    raise ValueError("CSV non leggibile (header mancante o delimitatore non supportato).")


def parse_float(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if s == "":
        return None
    s = s.replace(",", ".")
    try:
        return float(s)
    except Exception:
        return None


# ----------------------------
# Canonical conversion per modalità
# ----------------------------
def rows_to_canonical(mode: str, rows: list, solar_default_depth: float):
    """
    Ritorna lista di dict canonici:
    {
      id, rowId, description,
      headX,headY,headZ, endX,endY,endZ,
      diameter,
      treatment: { ... }   # solo per JET
    }
    """
    out = []

    if mode == "JET":
        # Jet template: PtNr, E/N/Z Head, E/N/Z End + H..AE treatment
        for r in rows:
            pid = r.get("PtNr")
            if _is_empty(pid):
                continue
            xh = parse_float(r.get("E Head"))
            yh = parse_float(r.get("N Head"))
            zh = parse_float(r.get("Z Head"))
            xe = parse_float(r.get("E End"))
            ye = parse_float(r.get("N End"))
            ze = parse_float(r.get("Z End"))
            if None in (xh, yh, zh, xe, ye, ze):
                continue

            treatment_keys = [
                "pr_1","drlStart_1","drlStop_1",
                "pr_2","drlStart_2","drlStop_2",
                "pr_3","drlStart_3","drlStop_3",
                "pr_4","drlStart_4","drlStop_4",
                "pr_j_1","jetStart_1","jetStop_1",
                "pr_j_2","jetStart_2","jetStop_2",
                "pr_j_3","jetStart_3","jetStop_3",
                "pr_j_4","jetStart_4","jetStop_4",
            ]
            treatment = {}
            for k in treatment_keys:
                v = r.get(k)
                if not _is_empty(v):
                    treatment[k] = str(v).strip()

            out.append({
                "id": str(pid).strip(),
                "rowId": None,
                "description": None,
                "headX": xh, "headY": yh, "headZ": zh,
                "endX": xe, "endY": ye, "endZ": ze,
                "diameter": None,
                "treatment": treatment
            })

    elif mode == "ROCK":
        # Rock template: RowNr, PtNr, head/end, Diameter, Description
        for r in rows:
            rownr = r.get("RowNr")
            pid = r.get("PtNr")
            if _is_empty(pid) and _is_empty(rownr):
                continue

            xh = parse_float(r.get("E Head"))
            yh = parse_float(r.get("N Head"))
            zh = parse_float(r.get("Z Head"))
            xe = parse_float(r.get("E End"))
            ye = parse_float(r.get("N End"))
            ze = parse_float(r.get("Z End"))
            if None in (xh, yh, zh, xe, ye, ze):
                continue

            dia = parse_float(r.get("Diameter"))
            desc = r.get("Description")
            desc = None if _is_empty(desc) else str(desc).strip()

            rid = None if _is_empty(rownr) else str(rownr).strip()
            pid_s = None if _is_empty(pid) else str(pid).strip()

            # HoleName: se ho RowNr e PtNr => "RowNr.PtNr" (compatibile col tuo parser)
            if rid and pid_s:
                hole_name = f"{rid}.{pid_s}"
            else:
                hole_name = pid_s or rid or "UNKNOWN"

            out.append({
                "id": hole_name,
                "rowId": rid,
                "description": desc,
                "headX": xh, "headY": yh, "headZ": zh,
                "endX": xe, "endY": ye, "endZ": ze,
                "diameter": dia,
                "treatment": {}  # niente jet treatment
            })

    elif mode == "SOLAR":
        # Solar template: PtNr, E/N/Z Head, Description
        # EndPoint calcolato: verticale con depth default
        for r in rows:
            pid = r.get("PtNr")
            if _is_empty(pid):
                continue

            xh = parse_float(r.get("E Head"))
            yh = parse_float(r.get("N Head"))
            zh = parse_float(r.get("Z Head"))
            if None in (xh, yh, zh):
                continue

            desc = r.get("Description")
            desc = None if _is_empty(desc) else str(desc).strip()

            xe, ye, ze = xh, yh, (zh + float(solar_default_depth))

            out.append({
                "id": str(pid).strip(),
                "rowId": None,
                "description": desc,
                "headX": xh, "headY": yh, "headZ": zh,
                "endX": xe, "endY": ye, "endZ": ze,
                "diameter": None,
                "treatment": {}  # niente jet treatment
            })
    else:
        raise ValueError(f"Unknown Mode: {mode}")

    return out


# ----------------------------
# IREDES Writer
# ----------------------------
def build_iredes_doc(plan_name: str, holes_canon: list, drillbit_dia_mm: int):
    root = ET.Element(f"{{{NS_DRP}}}DRPPlan", {
        "DRPPlanVersion": "V 1.0",
        "IRVersion": "V 1.0",
        "IRDownwCompat": "V 1.0",
        "DRPPlanDownwCompat": "V 1.0",
        f"xmlns:JET": NS_JET,
    })

    genhead = ET.SubElement(root, f"{{{NS_IR}}}GenHead")
    ET.SubElement(genhead, f"{{{NS_IR}}}FileCreateDate").text = (
        datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")
    )
    irver = ET.SubElement(genhead, f"{{{NS_IR}}}IRversion", {"DownwCompat": "V 1.0"})
    irver.text = "V 1.0"

    ET.SubElement(root, f"{{{NS_IR}}}PlanId").text = uuid.uuid4().hex[:12]
    ET.SubElement(root, f"{{{NS_IR}}}PlanName").text = plan_name
    ET.SubElement(root, f"{{{NS_IR}}}Comment").text = ""
    ET.SubElement(root, f"{{{NS_IR}}}Project").text = ""
    ET.SubElement(root, f"{{{NS_IR}}}WorkOrder").text = ""

    drillplan = ET.SubElement(root, "DrillPlan")
    ET.SubElement(drillplan, "NumberOfHoles").text = str(len(holes_canon))

    hole_id = 1
    for hc in holes_canon:
        hole = ET.SubElement(drillplan, "Hole")

        ET.SubElement(hole, "HoleId").text = str(hole_id)
        ET.SubElement(hole, "HoleName").text = str(hc["id"])

        if hc.get("description"):
            ET.SubElement(hole, "Comment").text = str(hc["description"])

        sp = ET.SubElement(hole, "StartPoint")
        ET.SubElement(sp, f"{{{NS_IR}}}PointX").text = f"{hc['headX']:.3f}"
        ET.SubElement(sp, f"{{{NS_IR}}}PointY").text = f"{hc['headY']:.3f}"
        ET.SubElement(sp, f"{{{NS_IR}}}PointZ").text = f"{hc['headZ']:.3f}"

        ep = ET.SubElement(hole, "EndPoint")
        ET.SubElement(ep, f"{{{NS_IR}}}PointX").text = f"{hc['endX']:.3f}"
        ET.SubElement(ep, f"{{{NS_IR}}}PointY").text = f"{hc['endY']:.3f}"
        ET.SubElement(ep, f"{{{NS_IR}}}PointZ").text = f"{hc['endZ']:.3f}"

        ET.SubElement(hole, "TypeOfHole").text = "Undefined"

        # Diameter: se presente nel ROCK, altrimenti default drillbit dia
        if hc.get("diameter") is not None:
            # Il tag standard nel tuo esempio è DrillBitDia: usiamo mm (se in input è mm già ok)
            ET.SubElement(hole, "DrillBitDia").text = str(int(round(hc["diameter"])))
        else:
            ET.SubElement(hole, "DrillBitDia").text = str(int(drillbit_dia_mm))

        # Custom geometry (utile ma non obbligatorio)
        g = compute_geometry(
            hc["headX"], hc["headY"], hc["headZ"],
            hc["endX"],  hc["endY"],  hc["endZ"]
        )
        jet_geom = ET.SubElement(hole, f"{{{NS_JET}}}Geometry")
        ET.SubElement(jet_geom, f"{{{NS_JET}}}Length").text = f"{g['length']:.3f}"
        ET.SubElement(jet_geom, f"{{{NS_JET}}}BearingDeg").text = f"{g['bearing_deg']:.3f}"
        ET.SubElement(jet_geom, f"{{{NS_JET}}}TiltDeg").text = f"{g['tilt_deg']:.3f}"
        ET.SubElement(jet_geom, f"{{{NS_JET}}}DeltaZ").text = f"{g['dz']:.3f}"

        # Custom treatment (solo se ci sono valori)
        treatment = hc.get("treatment") or {}
        if len(treatment) > 0:
            tr = ET.SubElement(hole, f"{{{NS_JET}}}Treatment")
            for k, v in treatment.items():
                # tag sicuro
                tag = "".join(ch if (ch.isalnum() or ch in ["_", "-"]) else "_" for ch in k.strip())
                ET.SubElement(tr, f"{{{NS_JET}}}{tag}").text = str(v)

        hole_id += 1

    tree = ET.ElementTree(root)
    ET.indent(tree, space="    ", level=0)
    return tree


def convert_file(mode: str, in_path: str, solar_default_depth: float, drillbit_dia_mm: int):
    ext = os.path.splitext(in_path)[1].lower()

    if ext in [".xlsx", ".xlsm", ".xls"]:
        headers, rows = read_table_from_excel(in_path)
    elif ext in [".csv"]:
        headers, rows = read_table_from_csv(in_path)
    else:
        raise ValueError("Unsupported file Format. Use .xlsx/.xlsm/.xls o .csv")

    holes_canon = rows_to_canonical(mode, rows, solar_default_depth=solar_default_depth)

    if len(holes_canon) == 0:
        raise ValueError("No valid Hole found (check column/values).")

    plan_name = os.path.splitext(os.path.basename(in_path))[0]
    tree = build_iredes_doc(plan_name, holes_canon, drillbit_dia_mm=drillbit_dia_mm)

    out_path = os.path.splitext(in_path)[0] + ".ird"
    tree.write(out_path, encoding="utf-8", xml_declaration=True)
    return out_path, len(holes_canon)


# ----------------------------
# GUI
# ----------------------------
def resource_path(relative):
    if getattr(sys, 'frozen', False):
        return os.path.join(sys._MEIPASS, relative)
    else:
        return os.path.join(os.path.abspath("."), relative)
class App(TkinterDnD.Tk):
    def __init__(self):
        super().__init__()
        self.title("STX Hole Converter")
        self.geometry("900x550")
# Imposta icona
      #  self.iconbitmap(resource_path("stx_icon.ico"))

        self.folders = ensure_project_folders()

        self.mode = None  # "ROCK" | "JET" | "SOLAR"

        # Top buttons
        top = tk.Frame(self)
        top.pack(fill="x", padx=12, pady=10)

        self.btn_rock = tk.Button(top, text="ROCK DRILL TEMPLATE", height=2, command=lambda: self.enter_mode("ROCK"))
        self.btn_jet  = tk.Button(top, text="JET GROUTING TEMPLATE", height=2, command=lambda: self.enter_mode("JET"))
        self.btn_solar= tk.Button(top, text="SOLAR FARM TEMPLATE", height=2, command=lambda: self.enter_mode("SOLAR"))

        self.btn_rock.pack(side="left", expand=True, fill="x", padx=6)
        self.btn_jet.pack(side="left",  expand=True, fill="x", padx=6)
        self.btn_solar.pack(side="left",expand=True, fill="x", padx=6)

        # Options
        opts = tk.LabelFrame(self, text="Opzioni")
        opts.pack(fill="x", padx=12, pady=(0, 10))

        self.var_drillbit = tk.StringVar(value="102")
        self.var_solar_depth = tk.StringVar(value="1.0")

        tk.Label(opts, text="DrillBitDia default (mm):").grid(row=0, column=0, sticky="w", padx=8, pady=6)
        tk.Entry(opts, textvariable=self.var_drillbit, width=10).grid(row=0, column=1, sticky="w", padx=8, pady=6)

        tk.Label(opts, text="Solar Offset default (m):").grid(row=0, column=2, sticky="w", padx=8, pady=6)
        self.entry_solar_depth = tk.Entry(opts, textvariable=self.var_solar_depth, width=10)
        self.entry_solar_depth.grid(row=0, column=3, sticky="w", padx=8, pady=6)

        tk.Label(opts, text="(Only SOLAR mode)").grid(row=0, column=4, sticky="w", padx=8, pady=6)

        # Drop area
        self.drop = tk.Label(
            self,
            text="Select a mode above.\nThen drag a file (xlsx/xlsm/csv) here\nor click to choose the file.",
            relief="groove",
            borderwidth=2,
            height=10,
            font=("Segoe UI", 11)
        )
        self.drop.pack(fill="both", expand=True, padx=12, pady=(0, 10))

        self.drop.drop_target_register(DND_FILES)
        self.drop.dnd_bind("<<Drop>>", self.on_drop)
        self.drop.bind("<Button-1>", self.on_click_pick)

        # Status
        self.status = tk.Label(self, text="Pronto.", anchor="w")
        self.status.pack(fill="x", padx=12, pady=(0, 12))

        self.refresh_ui_state()

    def refresh_ui_state(self):
        # entry solar depth abilitata solo in SOLAR
        if self.mode == "SOLAR":
            self.entry_solar_depth.config(state="normal")
        else:
            self.entry_solar_depth.config(state="disabled")

        if self.mode is None:
            self.drop.config(text="Select a mode above.\nThen drag a file (xlsx/xlsm/csv) here\nor click to choose the file.")
        else:
            self.drop.config(text=f"Mode:: {self.mode}\nDrag the file here (xlsx/xlsm/csv) to convert it in .ird\nOutput: same name, same folder.")

    def set_status(self, msg: str):
        self.status.config(text=msg)
        self.update_idletasks()

    def enter_mode(self, mode: str):
        self.mode = mode
        # crea cartelle e apre explorer nella sottocartella corretta
        self.folders = ensure_project_folders()
        open_in_explorer(self.folders[mode])
        self.set_status(f"Selected Mode: {mode} (folder opened in Explorer).")
        self.refresh_ui_state()

    def on_click_pick(self, event=None):
        if self.mode is None:
            messagebox.showwarning("Select Mode", "Choose ROCK / JET / SOLAR.")
            return
        path = filedialog.askopenfilename(
            initialdir=self.folders[self.mode],
            filetypes=[("Input files", "*.xlsx *.xlsm *.xls *.csv")]
        )
        if path:
            self.convert(path)

    def on_drop(self, event):
        if self.mode is None:
            messagebox.showwarning("Select Mode", "Choose ROCK / JET / SOLAR.")
            return

        files = self.tk.splitlist(event.data)
        if not files:
            return

        path = files[0]
        self.convert(path)


    def convert(self, in_path: str):
        try:
            if not os.path.isfile(in_path):
                raise FileNotFoundError("File not found.")

            drillbit = int(float(self.var_drillbit.get().strip().replace(",", ".")))
            solar_depth = float(self.var_solar_depth.get().strip().replace(",", "."))

            self.set_status(f"Converting ({self.mode}): {os.path.basename(in_path)} ...")
            out_path, n = convert_file(
                mode=self.mode,
                in_path=in_path,
                solar_default_depth=solar_depth,
                drillbit_dia_mm=drillbit
            )
            self.set_status(f"OK: created {os.path.basename(out_path)} ({n} holes)")
            messagebox.showinfo("Completed", f"File Created:\n{out_path}\nHoles: {n}")

        except Exception as e:
            self.set_status("Conversion Error.")
            messagebox.showerror("Error", str(e))


if __name__ == "__main__":
    App().mainloop()
